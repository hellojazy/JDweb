from __future__ import annotations

import copy
import json
import math
import re
import shutil
import sqlite3
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parent
UPLOAD_DIR = ROOT / "uploads"
OUTPUT_DIR = ROOT / "outputs"
HISTORY_FILE = OUTPUT_DIR / "history.json"
DB_FILE = ROOT / "jd_replenishment.sqlite3"

B_WAREHOUSE_TARGET_TURNOVER_DAYS = 14
DEFAULT_HOT_TURNOVER_DAYS = 30
DEFAULT_NORMAL_TURNOVER_DAYS = 25
DEFAULT_B_TURNOVER_DAYS = 14
DEFAULT_HOT_SALES_THRESHOLD = 1000
STAGNANT_SALES_14_THRESHOLD = 10
SKU_HEADERS = {"商品SKU", "商品编号", "商品编码", "SKU"}
PRODUCT_HEADERS = [
    "商品SKU",
    "商品名称",
    "品牌",
    "一二分类",
    "二级分类",
    "三级分类",
    "采控员",
    "采购员",
    "包装方案",
    "外箱箱规(X件/箱)",
    "是否畅销品",
]
PRODUCT_NAME_HEADERS = {"商品名称"}
BOX_SPEC_HEADERS = {"外箱箱规(X件/箱)"}


@dataclass
class ReplenishmentRow:
    sku: str
    center: str
    quantity: int
    box_spec: int
    supplier_code: str
    product_name: str
    band: str
    target_turnover: int
    near_14_daily_sales: float
    available_order_qty: float
    purchase_in_transit_qty: float
    source: str
    branch_gap: int = 0
    coverage_days: float | None = None

def normalize_sku(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def normalize_center(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("配送中心", "").strip()


def number(value: Any, default: float = 0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def positive_int(value: Any, default: int) -> int:
    try:
        parsed = int(float(value))
    except (TypeError, ValueError):
        return default
    return max(parsed, 0)


def excel_roundup(value: float) -> int:
    if value >= 0:
        return math.ceil(value)
    return math.floor(value)


def headers_for(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    return headers


def sku_column(headers: dict[str, int]) -> int | None:
    for name in SKU_HEADERS:
        if name in headers:
            return headers[name]
    return None


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)


def clear_data_rows(ws, start_row: int = 2) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_database() -> None:
    with db_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS products (
                sku TEXT PRIMARY KEY,
                data_json TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS generation_runs (
                run_id TEXT PRIMARY KEY,
                created_at TEXT NOT NULL,
                run_date TEXT NOT NULL,
                manual_count INTEGER NOT NULL,
                b_count INTEGER NOT NULL,
                manual_total_quantity INTEGER NOT NULL,
                b_total_quantity INTEGER NOT NULL,
                result_json TEXT NOT NULL
            )
            """
        )


def migrate_legacy_history_if_empty() -> None:
    init_database()
    with db_connection() as conn:
        existing = conn.execute("SELECT COUNT(*) FROM generation_runs").fetchone()[0]
    if existing or not HISTORY_FILE.exists():
        return
    try:
        records = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return
    if not isinstance(records, list):
        return
    for record in records:
        run_id = str(record.get("run_id") or "")
        result_path = OUTPUT_DIR / run_id / "result.json"
        if not result_path.exists():
            continue
        try:
            result = json.loads(result_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        add_history_record(result)


def product_value(row: dict[str, Any], names: set[str]) -> Any:
    for name in names:
        if name in row:
            return row.get(name)
    return None


def source_row_to_product(values: dict[str, Any], sku: str) -> dict[str, Any]:
    row = {header: values.get(header) for header in PRODUCT_HEADERS}
    row["商品SKU"] = sku
    if not row.get("商品名称"):
        row["商品名称"] = product_value(values, PRODUCT_NAME_HEADERS)
    if not row.get("外箱箱规(X件/箱)"):
        row["外箱箱规(X件/箱)"] = product_value(values, BOX_SPEC_HEADERS)
    return row


def read_product_metadata() -> dict[str, Any]:
    init_database()
    box_specs: dict[str, int] = {}
    product_by_sku: dict[str, str] = {}
    with db_connection() as conn:
        rows = conn.execute("SELECT sku, data_json FROM products ORDER BY rowid").fetchall()
    for item in rows:
        sku = normalize_sku(item["sku"])
        data = json.loads(item["data_json"])
        spec = number(data.get("外箱箱规(X件/箱)"))
        if sku and spec > 0:
            box_specs[sku] = int(spec)
        product_by_sku[sku] = str(data.get("商品名称") or "")
    return {"box_specs": box_specs, "product_by_sku": product_by_sku}


def list_products() -> dict[str, Any]:
    init_database()
    with db_connection() as conn:
        rows = conn.execute("SELECT sku, data_json FROM products ORDER BY rowid").fetchall()
    products: list[dict[str, Any]] = []
    for index, item in enumerate(rows, start=2):
        data = json.loads(item["data_json"])
        products.append({"row_number": index, "values": [data.get(header) for header in PRODUCT_HEADERS]})
    return {"headers": PRODUCT_HEADERS, "rows": products, "count": len(products)}


def import_products_from_excel(path: Path) -> dict[str, Any]:
    init_database()
    source_wb = openpyxl.load_workbook(path, data_only=True)
    source_ws = source_wb.active
    source_headers = headers_for(source_ws)
    source_sku_col = sku_column(source_headers)
    if not source_sku_col:
        raise ValueError("商品表缺少商品SKU列")

    with db_connection() as conn:
        old_skus = {row["sku"] for row in conn.execute("SELECT sku FROM products").fetchall()}

    skipped = 0
    source_by_sku: dict[str, dict[str, Any]] = {}
    source_order: list[str] = []
    for source_row in range(2, source_ws.max_row + 1):
        sku = normalize_sku(source_ws.cell(source_row, source_sku_col).value)
        if not sku:
            skipped += 1
            continue
        if sku not in source_by_sku:
            source_order.append(sku)
        values = {header: source_ws.cell(source_row, source_col).value for header, source_col in source_headers.items()}
        source_by_sku[sku] = source_row_to_product(values, sku)

    now = datetime.now().isoformat(timespec="seconds")
    with db_connection() as conn:
        conn.execute("DELETE FROM products")
        conn.executemany(
            "INSERT INTO products (sku, data_json, updated_at) VALUES (?, ?, ?)",
            [(sku, json.dumps(source_by_sku[sku], ensure_ascii=False), now) for sku in source_order],
        )

    products = list_products()
    new_skus = set(source_by_sku)
    return {
        "updated": len(old_skus & new_skus),
        "created": len(new_skus - old_skus),
        "deleted": len(old_skus - new_skus),
        "skipped": skipped,
        "total": products["count"],
        "products": products,
    }


def read_template_metadata() -> dict[str, Any]:
    product_meta = read_product_metadata()
    actual_b_warehouse = "成都补货B"

    return {
        "selected": set(),
        "box_specs": product_meta["box_specs"],
        "band_by_sku": {},
        "product_by_sku": product_meta["product_by_sku"],
        "supplier_by_sku": {},
        "actual_b_warehouse": actual_b_warehouse,
    }


def read_inventory(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    h = headers_for(ws)

    required = [
        "商品名称",
        "SKU",
        "供应商简码",
        "RDC",
        "配送中心",
        "可订购库存",
        "采购在途数量",
        "近14日出库商品件数",
        "14日有货天数",
    ]
    missing = [name for name in required if name not in h]
    if missing:
        raise ValueError(f"库存文件缺少列：{', '.join(missing)}")

    aggregated: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in range(2, ws.max_row + 1):
        sku = normalize_sku(ws.cell(row, h["SKU"]).value)
        rdc = normalize_center(ws.cell(row, h["RDC"]).value)
        center = normalize_center(ws.cell(row, h["配送中心"]).value)
        if not sku or not center:
            continue

        key = (sku, rdc, center)
        item = aggregated.setdefault(
            key,
            {
                "sku": sku,
                "rdc": rdc,
                "center": center,
                "product_name": ws.cell(row, h["商品名称"]).value or "",
                "supplier_code": ws.cell(row, h["供应商简码"]).value or "",
                "available_order_qty": 0.0,
                "purchase_in_transit_qty": 0.0,
                "sales_14": 0.0,
                "days_14": 0.0,
            },
        )
        item["available_order_qty"] += number(ws.cell(row, h["可订购库存"]).value)
        item["purchase_in_transit_qty"] += number(ws.cell(row, h["采购在途数量"]).value)
        item["sales_14"] += number(ws.cell(row, h["近14日出库商品件数"]).value)
        item["days_14"] = max(item["days_14"], number(ws.cell(row, h["14日有货天数"]).value))

    return list(aggregated.values())


def daily_sales(item: dict[str, Any]) -> float:
    sales_14 = number(item["sales_14"])
    days_14 = number(item["days_14"])
    return sales_14 / days_14 if days_14 else sales_14 / 14


def stock_coverage_days(available_qty: float, daily_qty: float) -> float | None:
    if daily_qty > 0:
        return available_qty / daily_qty
    return None


def classify_hot_skus(
    inventory: list[dict[str, Any]],
    hot_sales_threshold: int,
    actual_b_warehouse: str,
) -> dict[str, bool]:
    national_sales_by_sku: dict[str, float] = defaultdict(float)
    branch_sales_by_sku: dict[str, float] = defaultdict(float)

    for item in inventory:
        sku = item["sku"]
        center = item["center"]
        sales_14 = number(item["sales_14"])
        if center == "全国":
            national_sales_by_sku[sku] += sales_14
        elif center != actual_b_warehouse:
            branch_sales_by_sku[sku] += sales_14

    hot_by_sku: dict[str, bool] = {}
    for sku in set(national_sales_by_sku) | set(branch_sales_by_sku):
        sales_14 = national_sales_by_sku[sku] if sku in national_sales_by_sku else branch_sales_by_sku[sku]
        hot_by_sku[sku] = sales_14 > hot_sales_threshold
    return hot_by_sku


def calculate_b_warehouse_rows(
    inventory: list[dict[str, Any]],
    meta: dict[str, Any],
    manual_rows: list[ReplenishmentRow],
    branch_daily_by_sku: dict[str, float],
    b_turnover_days: int,
    hot_by_sku: dict[str, bool],
) -> list[ReplenishmentRow]:
    box_specs = meta["box_specs"]
    actual_b_warehouse = meta["actual_b_warehouse"]
    branch_gap_by_sku: dict[str, int] = defaultdict(int)
    manual_by_sku: dict[str, ReplenishmentRow] = {}

    for row in manual_rows:
        branch_gap_by_sku[row.sku] += row.quantity
        manual_by_sku.setdefault(row.sku, row)

    by_sku: dict[str, dict[str, Any]] = {}
    for item in inventory:
        if item["center"] != actual_b_warehouse:
            continue
        sku = item["sku"]
        bucket = by_sku.setdefault(
            sku,
            {
                "sku": sku,
                "available_order_qty": 0.0,
                "purchase_in_transit_qty": 0.0,
                "sales_14": 0.0,
                "days_14": 0.0,
                "supplier_code": item["supplier_code"],
                "product_name": item["product_name"],
            },
        )
        bucket["available_order_qty"] += number(item["available_order_qty"])
        bucket["purchase_in_transit_qty"] += number(item["purchase_in_transit_qty"])
        bucket["supplier_code"] = bucket["supplier_code"] or item["supplier_code"]
        bucket["product_name"] = bucket["product_name"] or item["product_name"]

    b_rows: list[ReplenishmentRow] = []
    for sku in sorted(set(by_sku) | set(branch_gap_by_sku) | set(branch_daily_by_sku)):
        item = by_sku.get(
            sku,
            {
                "sku": sku,
                "available_order_qty": 0.0,
                "purchase_in_transit_qty": 0.0,
                "sales_14": 0.0,
                "days_14": 0.0,
                "supplier_code": "",
                "product_name": "",
            },
        )
        box_spec = box_specs.get(sku)
        if not box_spec:
            continue

        daily_qty = branch_daily_by_sku.get(sku, 0.0)
        available_order_qty = number(item["available_order_qty"])
        purchase_in_transit_qty = number(item["purchase_in_transit_qty"])
        branch_gap = branch_gap_by_sku.get(sku, 0)
        target_qty = branch_gap + daily_qty * b_turnover_days
        quantity = excel_roundup((target_qty - available_order_qty - purchase_in_transit_qty) / box_spec) * box_spec

        if quantity > 0:
            manual_row = manual_by_sku.get(sku)
            b_rows.append(
                ReplenishmentRow(
                    sku=sku,
                    center=actual_b_warehouse,
                    quantity=int(quantity),
                    box_spec=int(box_spec),
                    supplier_code=str(
                        item["supplier_code"]
                        or (manual_row.supplier_code if manual_row else "")
                        or meta["supplier_by_sku"].get(sku, "")
                    ),
                    product_name=str(
                        meta["product_by_sku"].get(sku, "")
                        or item["product_name"]
                        or (manual_row.product_name if manual_row else "")
                    ),
                    band="热销品" if hot_by_sku.get(sku) else "普通品",
                    target_turnover=b_turnover_days,
                    near_14_daily_sales=daily_qty,
                    available_order_qty=available_order_qty,
                    purchase_in_transit_qty=purchase_in_transit_qty,
                    source="分仓差额+B仓周转",
                    branch_gap=branch_gap,
                    coverage_days=stock_coverage_days(available_order_qty, daily_qty),
                )
            )

    return sorted(b_rows, key=lambda row: row.sku)


def managed_product_name(item: dict[str, Any], product_by_sku: dict[str, str]) -> str:
    return str(product_by_sku.get(item["sku"]) or item.get("product_name") or "")


def build_inventory_warnings(
    inventory: list[dict[str, Any]],
    actual_b_warehouse: str,
    product_by_sku: dict[str, str],
) -> dict[str, list[dict[str, Any]]]:
    shortage: list[dict[str, Any]] = []
    stagnant: list[dict[str, Any]] = []

    for item in inventory:
        center = item["center"]
        if center == "全国":
            continue

        available_qty = number(item["available_order_qty"])
        daily_qty = daily_sales(item)
        coverage = stock_coverage_days(available_qty, daily_qty)

        if daily_qty > 0 and (coverage or 0) < B_WAREHOUSE_TARGET_TURNOVER_DAYS:
            shortage.append(
                {
                    "sku": item["sku"],
                    "product_name": managed_product_name(item, product_by_sku),
                    "rdc": item["rdc"],
                    "center": center,
                    "available_order_qty": available_qty,
                    "near_14_daily_sales": daily_qty,
                    "coverage_days": coverage,
                    "shortage_to_14_days": max(daily_qty * B_WAREHOUSE_TARGET_TURNOVER_DAYS - available_qty, 0),
                }
            )

        is_branch = center not in {"全国", actual_b_warehouse}
        if is_branch and available_qty > 0:
            sales_14 = number(item["sales_14"])
            if sales_14 < STAGNANT_SALES_14_THRESHOLD:
                stagnant.append(
                    {
                        "sku": item["sku"],
                        "product_name": managed_product_name(item, product_by_sku),
                        "rdc": item["rdc"],
                        "center": center,
                        "available_order_qty": available_qty,
                        "sales_14": sales_14,
                        "near_14_daily_sales": daily_qty,
                        "coverage_days": coverage,
                    }
                )

    shortage.sort(key=lambda row: row["shortage_to_14_days"], reverse=True)
    stagnant.sort(key=lambda row: (row["sales_14"], -row["available_order_qty"]))
    return {"shortage": shortage, "stagnant": stagnant}


def calculate_replenishment(
    inventory_path: Path,
    hot_turnover_days: int = DEFAULT_HOT_TURNOVER_DAYS,
    normal_turnover_days: int = DEFAULT_NORMAL_TURNOVER_DAYS,
    b_turnover_days: int = DEFAULT_B_TURNOVER_DAYS,
    hot_sales_threshold: int = DEFAULT_HOT_SALES_THRESHOLD,
) -> tuple[list[ReplenishmentRow], list[ReplenishmentRow], dict[str, Any], dict[str, list[dict[str, Any]]]]:
    meta = read_template_metadata()
    selected = meta["selected"]
    box_specs = meta["box_specs"]
    actual_b_warehouse = meta["actual_b_warehouse"]
    inventory = read_inventory(inventory_path)
    hot_by_sku = classify_hot_skus(inventory, hot_sales_threshold, actual_b_warehouse)

    manual_rows: list[ReplenishmentRow] = []
    branch_daily_by_sku: dict[str, float] = defaultdict(float)
    skipped_missing_box = 0
    skipped_not_selected = 0

    for item in inventory:
        sku = item["sku"]
        center = item["center"]
        if center in {"全国", actual_b_warehouse}:
            continue
        if selected and (sku, center) not in selected:
            skipped_not_selected += 1
            continue

        box_spec = box_specs.get(sku)
        if not box_spec:
            skipped_missing_box += 1
            continue

        is_hot = hot_by_sku.get(sku, False)
        band = "热销品" if is_hot else "普通品"
        target_turnover = hot_turnover_days if is_hot else normal_turnover_days
        daily_qty = daily_sales(item)
        branch_daily_by_sku[sku] += daily_qty
        available_order_qty = number(item["available_order_qty"])
        purchase_in_transit_qty = number(item["purchase_in_transit_qty"])
        raw_qty = (daily_qty * target_turnover - available_order_qty - purchase_in_transit_qty) / box_spec
        quantity = excel_roundup(raw_qty) * box_spec

        if quantity > 0:
            manual_rows.append(
                ReplenishmentRow(
                    sku=sku,
                    center=center,
                    quantity=int(quantity),
                    box_spec=int(box_spec),
                    supplier_code=str(item["supplier_code"] or meta["supplier_by_sku"].get(sku, "")),
                    product_name=str(meta["product_by_sku"].get(sku, "") or item["product_name"]),
                    band=band,
                    target_turnover=target_turnover,
                    near_14_daily_sales=daily_qty,
                    available_order_qty=available_order_qty,
                    purchase_in_transit_qty=purchase_in_transit_qty,
                    source="分仓补货",
                    coverage_days=stock_coverage_days(available_order_qty, daily_qty),
                )
            )

    manual_rows.sort(key=lambda row: (row.sku, row.center))
    b_rows = calculate_b_warehouse_rows(inventory, meta, manual_rows, branch_daily_by_sku, b_turnover_days, hot_by_sku)
    warnings = build_inventory_warnings(inventory, actual_b_warehouse, meta["product_by_sku"])
    summary = {
        "manual_count": len(manual_rows),
        "b_count": len(b_rows),
        "manual_total_quantity": sum(row.quantity for row in manual_rows),
        "b_total_quantity": sum(row.quantity for row in b_rows),
        "hot_turnover_days": hot_turnover_days,
        "normal_turnover_days": normal_turnover_days,
        "b_target_turnover": b_turnover_days,
        "hot_sales_threshold": hot_sales_threshold,
        "shortage_warning_count": len(warnings["shortage"]),
        "stagnant_warning_count": len(warnings["stagnant"]),
        "skipped_not_selected": skipped_not_selected,
        "skipped_missing_box": skipped_missing_box,
        "actual_b_warehouse": actual_b_warehouse,
    }
    return manual_rows, b_rows, summary, warnings


def prepare_output_sheet(ws, headers: list[str], widths: list[int]) -> None:
    ws.append(headers)
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="126F68")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="DCE5E4"))
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_index)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = widths[col_index - 1] if col_index <= len(widths) else 16
    ws.row_dimensions[1].height = 42


def fill_manual_workbook(rows: list[ReplenishmentRow], output_path: Path) -> None:
    headers = [
        "sku*",
        "采购渠道（格式：非渠道化）",
        "采购需求数量*",
        "供应商简码(非必填，不填则取默认供应商)",
        "期望下单日期范围(非必填，不填则立即自动下单)",
        "下单时间(格式：0、9、10...)",
        "配送中心*(格式：北京,上海,广州)",
        "是否需要供应商回告（格式：是、否，不填则默认是）",
        "备注",
        "建单是否允许部分成功（格式：是、否，不填则默认是）",
        "箱规（必填）",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "手工单"
    prepare_output_sheet(ws, headers, [18, 22, 16, 32, 32, 22, 28, 36, 18, 40, 14])
    for idx, row in enumerate(rows, start=2):
        ws.cell(idx, 1).value = row.sku
        ws.cell(idx, 3).value = row.quantity
        ws.cell(idx, 7).value = row.center
        ws.cell(idx, 11).value = row.box_spec
    wb.save(output_path)


def fill_b_workbook(rows: list[ReplenishmentRow], output_path: Path) -> None:
    headers = [
        "sku*",
        "采购渠道（格式：非渠道化、C采购渠道、B采购渠道，非必填，不填则系统默认获取）",
        "采购需求数量*",
        "供应商简码(采购类型为BBCC时必填)",
        "期望下单日期范围(格式：yyyymmdd-yyyymmdd非必填，不填则立即自动下单)",
        "下单时间(格式：0、9、10...21、22、23，非必填，不填则默认立即建单；当填写期望下单日期范围，下单时间必填)",
        "配送中心*(格式：可填写补货仓全选；可填写北京,上海,广州。如有复杂组合可填写补货仓全选,北京)",
        "币种（非必填，不填则默认RMB；全球购SKU建议填写）",
        "备注",
        "建单是否允许部分成功（格式：是、否，不填则默认是，仅允许全部留空或全部填是或全部填否；若已维护生单参数设置-商品维度拆单，全部填否即不允许部分成功的情况下已维护的商品维度拆单不生效）",
        "可替代老品（仅大件sku可用；格式：老品skuid，每行仅支持输入1个）",
        "sku起订量（仅大件sku可用；格式：0或正整数）",
        "是否校验箱规及箱规系数（格式：是、否，不填则默认是）",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    prepare_output_sheet(ws, headers, [18, 48, 16, 28, 42, 58, 52, 16, 18, 72, 34, 24, 30])
    for idx, row in enumerate(rows, start=2):
        ws.cell(idx, 1).value = row.sku
        ws.cell(idx, 3).value = row.quantity
        ws.cell(idx, 4).value = row.supplier_code or "zqsyq"
        ws.cell(idx, 7).value = row.center
        ws.cell(idx, 13).value = "是"
    wb.save(output_path)


def add_history_record(result: dict[str, Any]) -> None:
    init_database()
    summary = result["summary"]
    with db_connection() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO generation_runs (
                run_id, created_at, run_date, manual_count, b_count,
                manual_total_quantity, b_total_quantity, result_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                result["run_id"],
                result["created_at"],
                result["run_date"],
                summary["manual_count"],
                summary["b_count"],
                summary["manual_total_quantity"],
                summary["b_total_quantity"],
                json.dumps(result, ensure_ascii=False),
            ),
        )


def list_generation_history() -> list[dict[str, Any]]:
    init_database()
    migrate_legacy_history_if_empty()
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT run_id, created_at, run_date, manual_count, b_count,
                   manual_total_quantity, b_total_quantity, result_json
            FROM generation_runs
            ORDER BY created_at DESC
            """
        ).fetchall()
    records: list[dict[str, Any]] = []
    for row in rows:
        try:
            result = json.loads(row["result_json"])
        except json.JSONDecodeError:
            result = {}
        run_id = row["run_id"]
        if not (OUTPUT_DIR / run_id).exists():
            continue
        records.append(
            {
                "run_id": run_id,
                "created_at": row["created_at"],
                "run_date": row["run_date"],
                "manual_count": row["manual_count"],
                "b_count": row["b_count"],
                "manual_total_quantity": row["manual_total_quantity"],
                "b_total_quantity": row["b_total_quantity"],
                "files": result.get("files", {}),
            }
        )
    return records


def read_generation_record(run_id: str) -> dict[str, Any]:
    init_database()
    migrate_legacy_history_if_empty()
    with db_connection() as conn:
        row = conn.execute("SELECT result_json FROM generation_runs WHERE run_id = ?", (run_id,)).fetchone()
    if not row:
        raise FileNotFoundError(run_id)
    return json.loads(row["result_json"])


def generate_files(
    inventory_path: Path,
    run_date: date | None = None,
    hot_turnover_days: int = DEFAULT_HOT_TURNOVER_DAYS,
    normal_turnover_days: int = DEFAULT_NORMAL_TURNOVER_DAYS,
    b_turnover_days: int = DEFAULT_B_TURNOVER_DAYS,
    hot_sales_threshold: int = DEFAULT_HOT_SALES_THRESHOLD,
) -> dict[str, Any]:
    run_date = run_date or date.today()
    manual_rows, b_rows, summary, warnings = calculate_replenishment(
        inventory_path,
        hot_turnover_days=hot_turnover_days,
        normal_turnover_days=normal_turnover_days,
        b_turnover_days=b_turnover_days,
        hot_sales_threshold=hot_sales_threshold,
    )

    run_id = uuid4().hex[:12]
    output_dir = OUTPUT_DIR / run_id
    output_dir.mkdir(parents=True, exist_ok=True)

    date_text = run_date.strftime("%Y-%m-%d")
    time_text = datetime.now().strftime("%H%M%S")
    file_time_text = f"{date_text}-{time_text}"
    manual_name = f"手工单作业{file_time_text}.xlsx"
    b_name = f"京东入B仓{file_time_text}.xlsx"

    manual_path = output_dir / manual_name
    b_path = output_dir / b_name

    fill_manual_workbook(manual_rows, manual_path)
    fill_b_workbook(b_rows, b_path)

    result = {
        "run_id": run_id,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "run_date": date_text,
        "summary": summary,
        "manual_preview": [asdict(row) for row in manual_rows],
        "b_preview": [asdict(row) for row in b_rows],
        "warnings": warnings,
        "files": {
            "manual": {"name": manual_name, "url": f"/download/{run_id}/{manual_name}"},
            "b": {"name": b_name, "url": f"/download/{run_id}/{b_name}"},
        },
    }
    (output_dir / "result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    add_history_record(result)
    return result


def save_upload(filename: str, content: bytes) -> Path:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = Path(filename).name or "inventory.xlsx"
    target = UPLOAD_DIR / f"{uuid4().hex}_{safe_name}"
    target.write_bytes(content)
    return target


def reset_generated_outputs() -> None:
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
